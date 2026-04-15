#!/usr/bin/env python3
"""
bandcamp-push-log.py — Bandcamp inventory activity report.

Queries the last 8 weeks of:
  - channel_sync_log (Bandcamp inventory_push runs: completed, partial, paused)
  - warehouse_inventory_activity (source = 'bandcamp': sale events)

Outputs an Excel file with three sheets:
  - Push Runs   — every cron/triggered push run per workspace
  - Sale Events — every Bandcamp sale that triggered an inventory change
  - Timeline    — merged chronological view of both

Required env vars (same as .env.local):
  NEXT_PUBLIC_SUPABASE_URL
  SUPABASE_SERVICE_ROLE_KEY

Usage:
  python scripts/bandcamp-push-log.py
  python scripts/bandcamp-push-log.py --weeks 4
  python scripts/bandcamp-push-log.py --output /tmp/bandcamp-log.xlsx
"""

import argparse
import os
import sys
from datetime import datetime, timedelta, timezone

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl requests")
    sys.exit(1)

try:
    import requests
except ImportError:
    print("ERROR: requests not installed. Run: pip install openpyxl requests")
    sys.exit(1)


def get_env(key: str) -> str:
    val = os.environ.get(key, "")
    if not val:
        print(f"ERROR: {key} not set in environment.")
        sys.exit(1)
    return val


def supabase_get(url: str, service_key: str, path: str, params: dict) -> list:
    """Paginated Supabase REST query — fetches all rows."""
    headers = {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
        "Accept": "application/json",
        "Prefer": "count=exact",
    }
    rows = []
    limit = 1000
    offset = 0
    while True:
        p = {**params, "limit": limit, "offset": offset}
        r = requests.get(f"{url}/rest/v1/{path}", headers=headers, params=p, timeout=30)
        r.raise_for_status()
        batch = r.json()
        if not isinstance(batch, list):
            print(f"ERROR: unexpected response: {batch}")
            sys.exit(1)
        rows.extend(batch)
        if len(batch) < limit:
            break
        offset += limit
    return rows


def style_header(ws, row=1):
    fill = PatternFill("solid", fgColor="1E293B")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 60)


def fmt(ts: str | None) -> str:
    if not ts:
        return ""
    try:
        dt = datetime.fromisoformat(ts.replace("Z", "+00:00"))
        return dt.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    except Exception:
        return ts or ""


def main():
    parser = argparse.ArgumentParser(description="Bandcamp inventory push log")
    parser.add_argument("--weeks", type=int, default=8, help="Number of weeks to look back")
    parser.add_argument("--output", default="bandcamp-push-log.xlsx", help="Output file path")
    args = parser.parse_args()

    url = get_env("NEXT_PUBLIC_SUPABASE_URL")
    key = get_env("SUPABASE_SERVICE_ROLE_KEY")

    since = (datetime.now(timezone.utc) - timedelta(weeks=args.weeks)).isoformat()
    print(f"Fetching Bandcamp activity since {since} ({args.weeks} weeks)…")

    # Fetch push runs from channel_sync_log
    push_runs = supabase_get(
        url, key,
        "channel_sync_log",
        {
            "select": "id,workspace_id,channel,sync_type,status,items_processed,items_failed,started_at,completed_at,metadata",
            "channel": "eq.bandcamp",
            "sync_type": "eq.inventory_push",
            "completed_at": f"gte.{since}",
            "order": "completed_at.desc",
        },
    )
    print(f"  Push runs: {len(push_runs)}")

    # Fetch sale events from warehouse_inventory_activity
    sale_events = supabase_get(
        url, key,
        "warehouse_inventory_activity",
        {
            "select": "id,workspace_id,sku,delta,source,correlation_id,previous_quantity,new_quantity,metadata,created_at",
            "source": "eq.bandcamp",
            "created_at": f"gte.{since}",
            "is_synthetic": "eq.false",
            "order": "created_at.desc",
        },
    )
    print(f"  Sale events: {len(sale_events)}")

    wb = openpyxl.Workbook()

    # ── Sheet 1: Push Runs ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Push Runs"
    headers1 = [
        "Timestamp (UTC)", "Workspace ID", "Status",
        "Items Pushed", "Items Failed", "Run ID", "Band Count",
    ]
    ws1.append(headers1)
    style_header(ws1)

    for run in push_runs:
        meta = run.get("metadata") or {}
        ws1.append([
            fmt(run.get("completed_at")),
            run.get("workspace_id", ""),
            run.get("status", ""),
            run.get("items_processed", 0),
            run.get("items_failed", 0),
            meta.get("run_id", ""),
            meta.get("band_count", ""),
        ])
    autofit(ws1)

    # ── Sheet 2: Sale Events ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Sale Events")
    headers2 = [
        "Timestamp (UTC)", "Workspace ID", "SKU",
        "Delta", "Before", "After", "Correlation ID", "Bandcamp Item ID",
    ]
    ws2.append(headers2)
    style_header(ws2)

    for event in sale_events:
        meta = event.get("metadata") or {}
        ws2.append([
            fmt(event.get("created_at")),
            event.get("workspace_id", ""),
            event.get("sku", ""),
            event.get("delta", 0),
            event.get("previous_quantity", ""),
            event.get("new_quantity", ""),
            event.get("correlation_id", ""),
            meta.get("bandcamp_item_id", ""),
        ])
    autofit(ws2)

    # ── Sheet 3: Timeline (merged) ──────────────────────────────────────────
    ws3 = wb.create_sheet("Timeline")
    headers3 = ["Timestamp (UTC)", "Event Type", "Workspace ID", "Details"]
    ws3.append(headers3)
    style_header(ws3)

    timeline = []
    for run in push_runs:
        timeline.append((
            run.get("completed_at", ""),
            "Push Run",
            run.get("workspace_id", ""),
            f"Status: {run.get('status')} | {run.get('items_processed', 0)} pushed, {run.get('items_failed', 0)} failed",
        ))
    for event in sale_events:
        timeline.append((
            event.get("created_at", ""),
            "Bandcamp Sale",
            event.get("workspace_id", ""),
            f"SKU: {event.get('sku')} | delta: {event.get('delta', 0)} | {event.get('previous_quantity', '?')}→{event.get('new_quantity', '?')}",
        ))

    timeline.sort(key=lambda x: x[0] or "", reverse=True)
    for ts, event_type, ws_id, details in timeline:
        ws3.append([fmt(ts), event_type, ws_id, details])
    autofit(ws3)

    wb.save(args.output)
    print(f"\nSaved: {args.output}")
    print(f"  Sheet 'Push Runs':   {len(push_runs)} rows")
    print(f"  Sheet 'Sale Events': {len(sale_events)} rows")
    print(f"  Sheet 'Timeline':    {len(timeline)} rows")


if __name__ == "__main__":
    main()
