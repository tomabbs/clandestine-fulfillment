#!/usr/bin/env python3
"""
Image & Inventory Audit Spreadsheet
Tab 1: Per-product comparison (DB vs Shopify images + inventory)
Tab 2: Products in Shopify but not in DB, and DB products not in Shopify
"""

import json, math, sys, time
import urllib.request, urllib.error
from collections import defaultdict
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Credentials ──────────────────────────────────────────────────────────────
SUPABASE_URL  = "https://yspmgzphxlkcnfalndbh.supabase.co"
SUPABASE_KEY  = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
SHOPIFY_URL   = "https://kw16ph-t9.myshopify.com"
SHOPIFY_TOKEN = os.environ["SHOPIFY_ADMIN_API_TOKEN"]
SHOPIFY_VER   = "2026-01"

# ── Helpers ───────────────────────────────────────────────────────────────────

def sb_get(path, params=""):
    url = f"{SUPABASE_URL}/rest/v1/{path}?{params}" if params else f"{SUPABASE_URL}/rest/v1/{path}"
    req = urllib.request.Request(url, headers={
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Prefer": "count=exact",
    })
    with urllib.request.urlopen(req) as r:
        return json.loads(r.read())

def sb_get_all(path, select, filters="", page_size=1000):
    """Fetch all rows with pagination."""
    rows = []
    offset = 0
    while True:
        params = f"select={select}&limit={page_size}&offset={offset}"
        if filters:
            params += f"&{filters}"
        url = f"{SUPABASE_URL}/rest/v1/{path}?{params}"
        req = urllib.request.Request(url, headers={
            "apikey": SUPABASE_KEY,
            "Authorization": f"Bearer {SUPABASE_KEY}",
        })
        with urllib.request.urlopen(req) as r:
            batch = json.loads(r.read())
        rows.extend(batch)
        if len(batch) < page_size:
            break
        offset += page_size
        time.sleep(0.05)
    return rows

def shopify_gql(query, variables=None):
    endpoint = f"{SHOPIFY_URL}/admin/api/{SHOPIFY_VER}/graphql.json"
    payload = json.dumps({"query": query, "variables": variables or {}}).encode()
    req = urllib.request.Request(endpoint, data=payload, headers={
        "X-Shopify-Access-Token": SHOPIFY_TOKEN,
        "Content-Type": "application/json",
    })
    with urllib.request.urlopen(req) as r:
        return json.loads(r.read())

# ── Step 1: Fetch all DB products ─────────────────────────────────────────────
print("Fetching DB products...", flush=True)
db_products = sb_get_all(
    "warehouse_products",
    "id,title,shopify_product_id,status",
)
print(f"  {len(db_products)} products in DB")

# ── Step 2: Fetch all DB variants (SKU per product) ──────────────────────────
print("Fetching DB variants...", flush=True)
db_variants = sb_get_all(
    "warehouse_product_variants",
    "product_id,sku",
)
# Map product_id → list of SKUs
product_skus = defaultdict(list)
for v in db_variants:
    if v.get("sku"):
        product_skus[v["product_id"]].append(v["sku"])
print(f"  {len(db_variants)} variants")

# ── Step 3: DB image counts ───────────────────────────────────────────────────
print("Fetching DB image counts...", flush=True)
db_images = sb_get_all("warehouse_product_images", "product_id,shopify_image_id")
db_image_count   = defaultdict(int)   # total images in DB table
db_shopify_image_count = defaultdict(int)  # images with shopify_image_id (confirmed in Shopify)
for img in db_images:
    pid = img["product_id"]
    db_image_count[pid] += 1
    if img.get("shopify_image_id"):
        db_shopify_image_count[pid] += 1
print(f"  {len(db_images)} image rows")

# ── Step 4: DB inventory (by SKU) ────────────────────────────────────────────
print("Fetching DB inventory...", flush=True)
db_inventory = sb_get_all("warehouse_inventory_levels", "sku,available")
sku_inventory = {r["sku"]: r["available"] for r in db_inventory if r.get("sku")}

# ── Step 5: Bandcamp accounts per product ────────────────────────────────────
print("Fetching Bandcamp mappings...", flush=True)
bc_mappings = sb_get_all(
    "bandcamp_product_mappings",
    "variant_id,bandcamp_url",
    "bandcamp_url=not.is.null",
)
# Map variant_id → bandcamp subdomain
variant_bc_account = {}
for m in bc_mappings:
    url = m.get("bandcamp_url", "")
    # Extract subdomain from https://subdomain.bandcamp.com/...
    if "bandcamp.com" in url:
        try:
            subdomain = url.split("//")[1].split(".bandcamp.com")[0]
            variant_bc_account[m["variant_id"]] = subdomain
        except Exception:
            pass

# Map product_id → bandcamp account (take first match from any variant)
product_bc_account = {}
for v in db_variants:
    pid = v["product_id"]
    vid = v.get("id") or v.get("variant_id")
    if not vid:
        continue
    acc = variant_bc_account.get(vid)
    if acc and pid not in product_bc_account:
        product_bc_account[pid] = acc

# ── Step 6: Fetch all Shopify products with image + inventory counts ──────────
print("Fetching Shopify products (paginated)...", flush=True)

SHOPIFY_QUERY = """
query FetchProducts($first: Int!, $after: String) {
  products(first: $first, after: $after, sortKey: UPDATED_AT) {
    edges {
      node {
        id
        title
        status
        media(first: 50) { nodes { id } }
        variants(first: 10) {
          nodes {
            sku
            inventoryQuantity
          }
        }
      }
    }
    pageInfo { hasNextPage endCursor }
  }
}
"""

shopify_products = []  # list of {gid, title, status, image_count, total_inventory, skus}
cursor = None
page = 0
while True:
    page += 1
    print(f"  Shopify page {page} (fetched {len(shopify_products)} so far)...", flush=True)
    resp = shopify_gql(SHOPIFY_QUERY, {"first": 50, "after": cursor})
    edges = resp["data"]["products"]["edges"]
    page_info = resp["data"]["products"]["pageInfo"]
    for edge in edges:
        node = edge["node"]
        gid = node["id"]
        numeric_id = gid.split("/")[-1]
        image_count = len(node["media"]["nodes"])
        skus = [v["sku"] for v in node["variants"]["nodes"] if v.get("sku")]
        total_inv = sum(
            v.get("inventoryQuantity") or 0
            for v in node["variants"]["nodes"]
        )
        shopify_products.append({
            "gid": gid,
            "numeric_id": numeric_id,
            "title": node["title"],
            "status": node["status"],
            "image_count": image_count,
            "total_inventory": total_inv,
            "skus": skus,
        })
    if not page_info["hasNextPage"]:
        break
    cursor = page_info["endCursor"]
    time.sleep(0.3)  # gentle rate limiting

print(f"  {len(shopify_products)} products in Shopify")

# ── Step 7: Cross-reference ───────────────────────────────────────────────────
# DB lookup by shopify_product_id (numeric)
db_by_shopify_id = {}
db_by_id = {}
for p in db_products:
    db_by_id[p["id"]] = p
    sid = p.get("shopify_product_id")
    if sid:
        numeric = str(sid).split("/")[-1]  # normalise
        db_by_shopify_id[numeric] = p

# Shopify lookup by numeric ID
shopify_by_id = {sp["numeric_id"]: sp for sp in shopify_products}

# ── Step 8: Build Tab 1 rows ──────────────────────────────────────────────────
print("Building Tab 1...", flush=True)

# For each DB product that has a Shopify ID, merge data
tab1_rows = []
for p in db_products:
    sid = str(p.get("shopify_product_id") or "").split("/")[-1]
    sp = shopify_by_id.get(sid)

    skus = product_skus.get(p["id"], [])
    sku_str = ", ".join(skus)

    # DB inventory: sum over all SKUs
    db_inv = sum(sku_inventory.get(s, 0) for s in skus)

    # Shopify inventory
    shopify_inv = sp["total_inventory"] if sp else "Not in Shopify"
    shopify_img = sp["image_count"] if sp else "Not in Shopify"

    # DB images total
    db_img_total = db_image_count.get(p["id"], 0)
    # Images confirmed in Shopify (from sync)
    db_img_shopify = db_shopify_image_count.get(p["id"], 0)

    # Bandcamp account
    bc = product_bc_account.get(p["id"], "")

    tab1_rows.append({
        "Product Name": p["title"],
        "SKU(s)": sku_str,
        "Bandcamp Account": bc,
        "DB Status": p.get("status", ""),
        "Shopify Status": sp["status"].lower() if sp else "—",
        "# Images in DB": db_img_total,
        "# Images in Shopify": shopify_img,
        "Image Gap": (shopify_img - db_img_shopify) if isinstance(shopify_img, int) else "—",
        "Inventory in DB": db_inv,
        "Inventory in Shopify": shopify_inv,
    })

# Sort by image gap descending (products most missing images first)
def sort_key(r):
    gap = r["Image Gap"]
    return -(gap if isinstance(gap, (int, float)) else 0)
tab1_rows.sort(key=sort_key)

# ── Step 9: Build Tab 2 rows ──────────────────────────────────────────────────
print("Building Tab 2...", flush=True)

shopify_only = []   # In Shopify, NOT linked to any DB product
db_only      = []   # In DB with no shopify_product_id

for sp in shopify_products:
    if sp["numeric_id"] not in db_by_shopify_id:
        shopify_only.append({
            "Shopify Title": sp["title"],
            "Shopify Status": sp["status"],
            "Shopify ID": sp["numeric_id"],
            "Shopify SKUs": ", ".join(sp["skus"]),
            "Shopify Images": sp["image_count"],
            "Shopify Inventory": sp["total_inventory"],
        })

for p in db_products:
    sid = p.get("shopify_product_id")
    if not sid:
        skus = product_skus.get(p["id"], [])
        db_inv = sum(sku_inventory.get(s, 0) for s in skus)
        db_only.append({
            "DB Product Title": p["title"],
            "DB Status": p.get("status", ""),
            "DB SKU(s)": ", ".join(skus),
            "DB Images": db_image_count.get(p["id"], 0),
            "DB Inventory": db_inv,
        })

print(f"  {len(shopify_only)} in Shopify only (no DB link)")
print(f"  {len(db_only)} in DB only (no shopify_product_id)")

# ── Step 10: Write Excel ──────────────────────────────────────────────────────
print("Writing Excel file...", flush=True)

wb = openpyxl.Workbook()

HEADER_FILL  = PatternFill("solid", fgColor="1A1A2E")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL     = PatternFill("solid", fgColor="F5F5F5")
WARN_FILL    = PatternFill("solid", fgColor="FFF3CD")
BAD_FILL     = PatternFill("solid", fgColor="FFDDE0")
GOOD_FILL    = PatternFill("solid", fgColor="D4EDDA")
CENTER       = Alignment(horizontal="center", vertical="center")
LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin = Side(style="thin", color="DDDDDD")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, headers, col_widths):
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = col_widths[ci - 1]
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

# ── Tab 1: Product Comparison ─────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Product Comparison"

headers1 = [
    "Product Name", "SKU(s)", "Bandcamp Account",
    "DB Status", "Shopify Status",
    "# Images in DB", "# Images in Shopify", "Image Gap",
    "Inventory in DB", "Inventory in Shopify",
]
widths1 = [45, 20, 22, 12, 14, 14, 16, 12, 14, 16]
style_header(ws1, headers1, widths1)

for ri, row in enumerate(tab1_rows, 2):
    vals = [row[h] for h in headers1]
    for ci, val in enumerate(vals, 1):
        cell = ws1.cell(row=ri, column=ci, value=val)
        cell.border = BORDER
        # Alternate row shading
        if ri % 2 == 0:
            cell.fill = ALT_FILL

    # Highlight rows where Shopify has images but DB images in Shopify column is low
    img_shopify = row["# Images in Shopify"]
    img_db      = row["# Images in DB"]
    gap         = row["Image Gap"]

    if isinstance(img_shopify, int) and img_shopify == 0:
        # No images in Shopify at all → red
        for ci in range(6, 9):
            ws1.cell(row=ri, column=ci).fill = BAD_FILL
    elif isinstance(gap, int) and gap > 0:
        # Shopify has more than DB confirms → warning
        for ci in range(6, 9):
            ws1.cell(row=ri, column=ci).fill = WARN_FILL
    elif img_db > 0 and isinstance(img_shopify, int) and img_shopify > 0:
        # Both have images
        for ci in range(6, 9):
            ws1.cell(row=ri, column=ci).fill = GOOD_FILL

    # Center numeric cols
    for ci in range(6, 11):
        ws1.cell(row=ri, column=ci).alignment = CENTER
    # Wrap text for name/sku cols
    ws1.cell(row=ri, column=1).alignment = LEFT
    ws1.cell(row=ri, column=2).alignment = LEFT

# Add summary row
total_row = len(tab1_rows) + 3
ws1.cell(row=total_row, column=1, value=f"Total: {len(tab1_rows)} products")
ws1.cell(row=total_row, column=1).font = Font(bold=True)

no_shopify_imgs = sum(1 for r in tab1_rows if r["# Images in Shopify"] == 0)
ws1.cell(row=total_row+1, column=1, value=f"Products with 0 images in Shopify: {no_shopify_imgs}")
ws1.cell(row=total_row+1, column=1).font = Font(bold=True, color="CC0000")

# ── Tab 2: Shopify Only ────────────────────────────────────────────────────────
ws2 = wb.create_sheet("In Shopify Not in DB")
headers2 = ["Shopify Title", "Shopify Status", "Shopify ID", "Shopify SKUs", "Shopify Images", "Shopify Inventory"]
widths2 = [50, 14, 18, 30, 14, 16]
style_header(ws2, headers2, widths2)
for ri, row in enumerate(sorted(shopify_only, key=lambda x: x["Shopify Title"]), 2):
    for ci, h in enumerate(headers2, 1):
        cell = ws2.cell(row=ri, column=ci, value=row[h])
        cell.border = BORDER
        if ri % 2 == 0:
            cell.fill = ALT_FILL
    ws2.cell(row=ri, column=1).alignment = LEFT
ws2.cell(row=len(shopify_only)+3, column=1, value=f"Total: {len(shopify_only)} Shopify-only products").font = Font(bold=True)

# ── Tab 3: DB Only ────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("In DB Not in Shopify")
headers3 = ["DB Product Title", "DB Status", "DB SKU(s)", "DB Images", "DB Inventory"]
widths3 = [50, 12, 30, 12, 14]
style_header(ws3, headers3, widths3)
for ri, row in enumerate(sorted(db_only, key=lambda x: x["DB Product Title"]), 2):
    for ci, h in enumerate(headers3, 1):
        cell = ws3.cell(row=ri, column=ci, value=row[h])
        cell.border = BORDER
        if ri % 2 == 0:
            cell.fill = ALT_FILL
    ws3.cell(row=ri, column=1).alignment = LEFT
ws3.cell(row=len(db_only)+3, column=1, value=f"Total: {len(db_only)} DB-only products").font = Font(bold=True)

# ── Save ──────────────────────────────────────────────────────────────────────
out = "/Users/Shared/WorkShared/Project/clandestine-fulfillment/docs/image-inventory-audit.xlsx"
wb.save(out)
print(f"\nDone! Saved to: {out}")
print(f"\nSummary:")
print(f"  Tab 1 - Product Comparison: {len(tab1_rows)} rows")
print(f"  Tab 2 - In Shopify not in DB: {len(shopify_only)} rows")
print(f"  Tab 3 - In DB not in Shopify: {len(db_only)} rows")
print(f"  Products with 0 images in Shopify: {no_shopify_imgs}")
